import openpyxl
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from prettytable import PrettyTable
from codings import config as cfg
from sklearn.metrics import roc_curve, auc


def classification_res():
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [('Method', "Accuracy (%)", "Precision (%)", "Recall (%)", "F1-score (%)", "Specificity (%)",
              "FPR", "FNR"),
            ("Proposed", cfg.pr_acc, cfg.pr_pre, cfg.pr_rec, cfg.pr_fsc, cfg.pr_spec, cfg.pr_fpr, cfg.pr_fnr),
            ("BNN", cfg.bnn_acc, cfg.bnn_pre, cfg.bnn_rec, cfg.bnn_fsc, cfg.bnn_spec, cfg.bnn_fpr, cfg.bnn_fnr),
            ("CAM", cfg.cam_acc, cfg.cam_pre, cfg.cam_rec, cfg.cam_fsc, cfg.cam_spec, cfg.cam_fpr, cfg.cam_fnr),
            ("CNN", cfg.cnn_acc, cfg.cnn_pre, cfg.cnn_rec, cfg.cnn_fsc, cfg.cnn_spec, cfg.cnn_fpr, cfg.cnn_fnr),
            ("LSTM", cfg.lstm_acc, cfg.lstm_pre, cfg.lstm_rec, cfg.lstm_fsc, cfg.lstm_spec, cfg.lstm_fpr, cfg.lstm_fnr)]
    for row in rows:
        ws.append(row)
    wb.save("..\\Results\\classifier_res.xlsx")
def accuracy():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'BNN', 'CAM', 'CNN', 'LSTM']
    time = [cfg.pr_acc, cfg.bnn_acc, cfg.cam_acc, cfg.cnn_acc, cfg.lstm_acc]
    X_axis = np.arange(len(X))
    plt.bar(X, time, color="lightcoral")
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("Accuracy (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\accuracy.png")
    plt.show()
def precision():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'BNN', 'CAM', 'CNN', 'LSTM']
    time = [cfg.pr_pre, cfg.bnn_pre, cfg.cam_pre, cfg.cnn_pre, cfg.lstm_pre]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color="green", marker="o", ls="--", markeredgecolor='black')
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    # plt.ylim(0, 0.1)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("Precision (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\precision.png")
    plt.show()
def recall():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'BNN', 'CAM', 'CNN', 'LSTM']
    time = [cfg.pr_rec, cfg.bnn_rec, cfg.cam_rec, cfg.cnn_rec, cfg.lstm_rec]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color="teal", marker="o", linestyle="dashdot", markeredgecolor='black')
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("Recall (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\recall.png")
    plt.show()
def fmeasure():
    plt.figure(figsize=(8, 6))
    X = ['Proposed', 'BNN', 'CAM', 'CNN', 'LSTM']
    time = [cfg.pr_fsc, cfg.bnn_fsc, cfg.cam_fsc, cfg.cnn_fsc, cfg.lstm_fsc]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color='plum', marker='o')  # You can adjust the color and marker style as needed
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    # plt.ylim(0, 0.1)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("F-measure (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\fm.png")
    plt.show()
def specificity():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'BNN', 'CAM', 'CNN', 'LSTM']
    colors = ['wheat']
    time = [cfg.pr_spec, cfg.bnn_spec, cfg.cam_spec, cfg.cnn_spec, cfg.lstm_spec]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color='orange', marker='o')  # You can adjust the color and marker style as needed
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    # plt.ylim(0, 0.1)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("Specificity (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\specificity.png")
    plt.show()
def fpr():
    plt.figure(figsize=(8, 6))
    X = ['Proposed', 'BNN', 'CAM', 'CNN', 'LSTM']
    colors = ['darksalmon']
    time = [cfg.pr_fpr, cfg.bnn_fpr, cfg.cam_fpr, cfg.cnn_fpr, cfg.lstm_fpr]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color='red', marker='o')  # You can adjust the color and marker style as needed
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    # plt.ylim(0, 0.1)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("FPR", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\FPR.png")
    plt.show()
def fnr():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'BNN', 'CAM', 'CNN', 'LSTM']
    colors = ['wheat', 'darksalmon', 'plum', 'y', 'darkturquoise', 'orange', 'paleturquoise']
    time = [cfg.pr_fnr, cfg.bnn_fnr, cfg.cam_fnr, cfg.cnn_fnr, cfg.lstm_fnr]
    X_axis = np.arange(len(X))
    plt.bar(X, time, color=colors)
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    # plt.ylim(0, 0.1)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("FNR", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\FNR.png")
    plt.show()

def risk_res():
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [('Method', "Accuracy (%)", "Precision (%)", "Recall (%)", "F1-score (%)"),
            ("Pneumonia Risk Scoring and Estimation (P-RiSE)", cfg.prise_acc, cfg.prise_pre, cfg.prise_rec, cfg.prise_fsc),
            ("Fuzzy Rule", cfg.fr_acc, cfg.fr_pre, cfg.fr_rec, cfg.fr_fsc),
            ("Deep Neural Network (DNN)", cfg.dnn_acc, cfg.dnn_pre, cfg.dnn_rec, cfg.dnn_fsc),
            ("Artificial Neural Network (ANN)", cfg.ann_acc, cfg.ann_pre, cfg.ann_rec, cfg.ann_fsc),
            ("Feed Forward Network (FFN)", cfg.fnn_acc, cfg.fnn_pre, cfg.fnn_rec, cfg.fnn_fsc)]
    for row in rows:
        ws.append(row)
    wb.save("..\\Results\\Risk_res.xlsx")
def risk_accuracy():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'Fuzzy Rule', 'DNN', 'ANN', 'FFN']
    time = [cfg.prise_acc, cfg.fr_acc, cfg.dnn_acc, cfg.ann_acc, cfg.fnn_acc]
    X_axis = np.arange(len(X))
    plt.bar(X, time, color="green")
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("Accuracy (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\risk_accuracy.png")
    plt.show()
def risk_precision():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'Fuzzy Rule', 'DNN', 'ANN', 'FFN']
    time = [cfg.prise_pre, cfg.fr_pre, cfg.dnn_pre, cfg.ann_pre, cfg.fnn_pre]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color="brown", marker='o')
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("Precision (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\risk_precision.png")
    plt.show()
def risk_recall():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'Fuzzy Rule', 'DNN', 'ANN', 'FFN']
    time = [cfg.prise_rec, cfg.fr_rec, cfg.dnn_rec, cfg.ann_rec, cfg.fnn_rec]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color="orangered", marker='o')
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("Recall (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\risk_recall.png")
    plt.show()
def risk_fsc():
    plt.figure(figsize=(8, 6))

    X = ['Proposed', 'Fuzzy Rule', 'DNN', 'ANN', 'FFN']
    time = [cfg.prise_fsc, cfg.fr_fsc, cfg.dnn_fsc, cfg.ann_fsc, cfg.fnn_fsc]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color="purple", marker='o')
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("F-Score (%)", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\risk_fscore.png")
    plt.show()

def fusion_res():
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = [('Method', "MSE", "MAE"),
            ("Proposed OTFN", cfg.pr_mse, cfg.pr_mae),
            ("TFN", cfg.tfn_mse, cfg.tfn_mae),
            ("Low Rank RFN (LRTFN)", cfg.lrtfn_mse, cfg.lrtfn_mae),
            ("Late Fusion (LF)", cfg.lf_mse, cfg.lf_mae),
            ("Early Fusion (EF)", cfg.ef_mse, cfg.ef_mae)]
    for row in rows:
        ws.append(row)
    wb.save("..\\Results\\Fusion_res.xlsx")
def mae():
    plt.figure(figsize=(8, 6))

    X = ['Proposed OTFN', 'TFN', 'LRTFN', 'LF', 'EF']
    time = [cfg.pr_mae, cfg.tfn_mae, cfg.lrtfn_mae, cfg.lf_mae, cfg.ef_mae]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color='cyan', marker='o')  # You can adjust the color and marker style as needed
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    # plt.ylim(0, 0.1)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("MAE", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\MAE.png")
    plt.show()
def mse():
    plt.figure(figsize=(8, 6))

    X = ['Proposed OTFN', 'TFN', 'LRTFN', 'LF', 'EF']
    time = [cfg.pr_mse, cfg.tfn_mse, cfg.lrtfn_mse, cfg.lf_mse, cfg.ef_mse]
    X_axis = np.arange(len(X))
    plt.plot(X, time, color='dodgerblue', marker='o')  # You can adjust the color and marker style as needed
    plt.xticks(X_axis, X, font="Times New Roman", fontweight='bold', fontsize=12)
    # plt.ylim(0, 0.1)
    plt.yticks(font="Times New Roman", fontweight='bold', fontsize=12)
    plt.xlabel("Techniques", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.ylabel("MSE", font="Times New Roman", fontweight="bold", fontsize=12)
    plt.savefig("..\\Results\\MSE.png")
    plt.show()


